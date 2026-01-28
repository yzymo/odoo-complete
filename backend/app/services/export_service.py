"""
Export service for generating Excel files from product data.
"""

import logging
from typing import List, Dict, Any
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.config import get_storage_path

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting products to Excel format."""

    # Colonnes Excel avec tous les champs Odoo
    EXCEL_COLUMNS = [
        # Identifiants
        ("default_code", "Référence Interne"),
        ("barcode", "Code-barres"),
        ("Code_EAN", "Code EAN"),

        # Informations Produit
        ("name", "Nom du Produit"),
        ("type", "Type"),
        ("active", "Actif"),
        ("is_published", "Publié"),

        # Classification
        ("categ_id", "Catégorie"),
        ("country_of_origin", "Pays d'Origine"),

        # Fabricant
        ("constructeur", "Fabricant"),
        ("refConstructeur", "Réf. Fabricant"),

        # Descriptions
        ("description_courte", "Description Courte"),
        ("description_ecommerce", "Description E-commerce"),
        ("features_description", "Caractéristiques"),

        # Dimensions
        ("length", "Longueur (mm)"),
        ("width", "Largeur (mm)"),
        ("height", "Hauteur (mm)"),
        ("weight", "Poids (kg)"),

        # Logistique
        ("hs_code", "Code Douanier"),
        ("contient_du_lithium", "Contient Lithium"),

        # Prix
        ("lst_price", "Prix Catalogue"),
        ("taxes_id", "Taxes"),

        # Images
        ("image_count", "Nombre d'Images"),
        ("image_512", "Image Principale"),

        # Documents
        ("fiche_constructeur_nom", "Fiche Constructeur"),
        ("fiche_technique_nom", "Fiche Technique"),

        # Métadonnées Extraction
        ("extraction_status", "Statut"),
        ("extraction_confidence", "Score Confiance"),
        ("source_files", "Fichiers Sources"),

        # Dates
        ("created_at", "Date Création"),
        ("updated_at", "Date Modification"),
    ]

    def __init__(self):
        """Initialize export service with storage path."""
        self.export_dir = get_storage_path("exports")

    def create_excel_file(
        self,
        products: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        Create Excel file from products list.

        Args:
            products: List of product dicts
            filename: Optional filename (auto-generated if not provided)

        Returns:
            Path to created Excel file
        """
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                filename = f"products_export_{timestamp}.xlsx"

            file_path = os.path.join(self.export_dir, filename)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Produits"

            # Write header row
            self._write_header(ws)

            # Write product rows
            for row_idx, product in enumerate(products, start=2):
                self._write_product_row(ws, row_idx, product)

            # Auto-adjust column widths
            self._adjust_column_widths(ws)

            # Save workbook
            wb.save(file_path)

            logger.info(f"Excel file created: {file_path} ({len(products)} products)")
            return file_path

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            raise

    def _write_header(self, ws):
        """Write header row with styling."""
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (field_key, field_label) in enumerate(self.EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field_label
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _write_product_row(self, ws, row_idx: int, product: Dict[str, Any]):
        """Write a single product row."""
        for col_idx, (field_key, field_label) in enumerate(self.EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = self._get_field_value(product, field_key)
            cell.value = value

            # Alignment
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, bool):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    def _get_field_value(self, product: Dict[str, Any], field_key: str) -> Any:
        """Extract and format field value from product."""
        # Direct fields
        if field_key in product:
            value = product[field_key]

            # Format special types
            if isinstance(value, bool):
                return "Oui" if value else "Non"
            elif isinstance(value, list):
                if field_key == "taxes_id":
                    return ", ".join(value) if value else ""
                return str(len(value)) if value else "0"
            elif isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                return ""
            else:
                return value

        # Computed fields
        if field_key == "image_count":
            return len(product.get("images", []))

        if field_key == "extraction_status":
            metadata = product.get("extraction_metadata", {})
            return metadata.get("status", "unknown")

        if field_key == "extraction_confidence":
            metadata = product.get("extraction_metadata", {})
            scores = metadata.get("field_confidence_scores", {})
            if scores:
                avg_score = sum(scores.values()) / len(scores)
                return round(avg_score * 100, 1)  # Percentage
            return ""

        if field_key == "source_files":
            sources = product.get("sources", [])
            if sources:
                filenames = [s.get("origin_file", "") for s in sources]
                return ", ".join(set(filenames))  # Unique filenames
            return ""

        return ""

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""
        for col_idx, (field_key, field_label) in enumerate(self.EXCEL_COLUMNS, start=1):
            column_letter = get_column_letter(col_idx)

            # Set minimum width based on header
            max_length = len(field_label)

            # Check a few rows to determine width (not all for performance)
            for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row), min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            # Set width with limits
            adjusted_width = min(max_length + 2, 50)  # Max 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_excel_with_filters(
        self,
        products: List[Dict[str, Any]],
        filters: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Create Excel file with filter info.

        Args:
            products: List of product dicts
            filters: Applied filters info
            filename: Optional filename

        Returns:
            Path to created Excel file
        """
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                filename = f"products_export_{timestamp}.xlsx"

            file_path = os.path.join(self.export_dir, filename)

            # Create workbook
            wb = Workbook()

            # Products sheet
            ws_products = wb.active
            ws_products.title = "Produits"
            self._write_header(ws_products)

            for row_idx, product in enumerate(products, start=2):
                self._write_product_row(ws_products, row_idx, product)

            self._adjust_column_widths(ws_products)

            # Info sheet
            ws_info = wb.create_sheet("Informations Export")
            self._write_info_sheet(ws_info, products, filters)

            # Save workbook
            wb.save(file_path)

            logger.info(f"Excel file with filters created: {file_path}")
            return file_path

        except Exception as e:
            logger.error(f"Error creating Excel file with filters: {e}")
            raise

    def _write_info_sheet(
        self,
        ws,
        products: List[Dict[str, Any]],
        filters: Dict[str, Any]
    ):
        """Write export information sheet."""
        # Title
        ws.cell(row=1, column=1, value="Informations sur l'Export")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        row = 3

        # Export info
        ws.cell(row=row, column=1, value="Date d'export:")
        ws.cell(row=row, column=2, value=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
        row += 1

        ws.cell(row=row, column=1, value="Nombre de produits:")
        ws.cell(row=row, column=2, value=len(products))
        row += 2

        # Filters applied
        if filters:
            ws.cell(row=row, column=1, value="Filtres appliqués:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            for filter_key, filter_value in filters.items():
                ws.cell(row=row, column=1, value=f"  {filter_key}:")
                ws.cell(row=row, column=2, value=str(filter_value))
                row += 1

        # Statistics
        row += 1
        ws.cell(row=row, column=1, value="Statistiques:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Count by status
        status_counts = {}
        for product in products:
            status = product.get("extraction_metadata", {}).get("status", "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1

        for status, count in status_counts.items():
            ws.cell(row=row, column=1, value=f"  {status}:")
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Adjust widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
